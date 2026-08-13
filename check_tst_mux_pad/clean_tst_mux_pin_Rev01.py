import re
import time
import openpyxl
from openpyxl.styles import Font
import pandas as pd
from JonathanPyLib_V13 import *


INPUT_FOLDER = './input/'
OUTPUT_FOLDER = './output/'
SPACE_NUMBER:int = 4


class CleanTstMuxPin():
    
    def __init__(self, tstfile, pin_mux_alias_excel, extention):
        self.tstfile = tstfile
        self.tst_name = str(tstfile.split('/')[-1]).strip('.tst.gz')
        self.pin_mux_alias_excel =  pin_mux_alias_excel
        self.tst_RW = ReadWriteFiles()
        self.extention = extention
        if extention == '.gz':
            self.tst_RW.read_gz_lines(self.tstfile, 'rt')       # read tst.gz or .tst file content
        elif extention == '.tst':
            self.tst_RW.read_file_lines(self.tstfile)
        
        self.tst_content = self.tst_RW.read_content
        
        self.mux_group = {}                                 # mux_group['PAD_BRCS0'] = ['PAD_BRCS0', 'PAD_MOD_TX_N00'...] from excel
        self.establish_pin_mux_group()                      # establish pin mux group dictionary from channel map
        
        self.all_digital_pads = []                          # all used digital pads in alias excel
        self.get_all_digital_pads()
        
        self.tst_assign_pads = []                           # ASSIGN pads in tst
        self.alias = {}                                     # self.alias['/* 1 */'] = ['X', 'X', '0', 'H', ...]
        self.extract_assign_pads_and_alias()
        
        self.tst_dataframe = pd.DataFrame()                 # row: vector number of tst, column: ASSIGN pads
        self.create_pad_alias_dataframe()
        
        self.all_mux_pads_in_this_tst = []                  # Mux pads used in tst
        self.mux_X_pads_in_this_tst = []                    # don't care Mux pads used in tst
        self.pads_not_exist_in_excel_but_in_tst = []        # pads not exist in excel but in tst
        self.remove_X_pinmux_pads_in_dataframe()
        
        self.new_alias = {}                                 # self.new_alias['/* 1 */'] = ['0', '1', '0', 'H', ...]
        self.write_pinmux_cleaned_tst()
        
        
    def extract_assign_pads_and_alias(self):
        i=0
        assign_detect_flag = False
        for line in self.tst_content:
            if(re.match(r"^ASSIGN.+,\n", line) and not assign_detect_flag):
                assign_detect_flag = True
                line=line.strip('ASSIGN').strip().strip(',')
                self.tst_assign_pads = line.split(',')
            if(re.match(r"^\s+PAD_.+,\n", line) and assign_detect_flag):
                line=line.strip().strip(',')
                self.tst_assign_pads.extend(line.split(','))
            if(re.match(r"^\s+PAD_.+;\n", line) and assign_detect_flag):
                line=line.strip().strip(';')
                self.tst_assign_pads.extend(line.split(','))
            if(re.match(r"\w+;\s\/\*\s\d+\s\*\/", line)):
                alias_key = line.split(';')[1]
                separate_aliases = list(line.split(';')[0])
                
                self.alias[alias_key] = separate_aliases
                i+=1
                # if(i==8000):
                #     break
        print(" "*SPACE_NUMBER + "Total ASSIGN pads unmber in tst: ", len(self.tst_assign_pads))

        
    def establish_pin_mux_group(self):
        pin_mux_df = pd.read_excel(self.pin_mux_alias_excel, sheet_name='ALIAS')
        for row in pin_mux_df.loc[:, 'MUX0']:
            if (str(row) != 'nan'):
                self.mux_group[row] = [str(row)]
        for i in range(len(pin_mux_df.loc[:, 'MUX0'])):
            for mux in pin_mux_df.columns:
                if(mux != 'MUX0' and re.match(r"MUX\d+", mux, re.IGNORECASE) and str(pin_mux_df.loc[i, mux]) != 'nan'):
                    self.mux_group[pin_mux_df.loc[i, 'MUX0']].append(str(pin_mux_df.loc[i, mux]))
        # dump_dic(self.mux_group)
    
    
    def get_all_digital_pads(self):
        all_digital_pads_df = pd.read_excel(self.pin_mux_alias_excel, sheet_name='ALIAS')
        for row in range(len(all_digital_pads_df.iloc[:, 0])):
            for col in range(len(all_digital_pads_df.columns)):
                if (str(all_digital_pads_df.iloc[row, col]) != 'nan'):
                    self.all_digital_pads.append(str(all_digital_pads_df.iloc[row, col]))
        print(" "*SPACE_NUMBER + "All assigned digital & mux pads number in \"" + str(self.pin_mux_alias_excel).split('/')[-1] +"\":", len(self.all_digital_pads))
        
    
    def create_pad_alias_dataframe(self):
        data = pd.DataFrame()
        for col in range(0, len(self.tst_assign_pads)):
            col_list_alias = []
            for row in self.alias:
                col_list_alias.append(self.alias[row][col])
            Ser = pd.Series(col_list_alias, index = range(1, len(self.alias)+1))    #Ser1 = pd.Series(range(0,7), index = range(100, 107))
            data_each = pd.DataFrame({ f'{self.tst_assign_pads[col]}' : Ser }) 
            data = pd.concat([data, data_each], axis=1)
            progress = col/(len(self.tst_assign_pads)-1)*100
            sys.stdout.write("\r" + " "*SPACE_NUMBER + "Building tst content dataframe from original -> "+ "{:.2f}".format(progress) +"%")
            sys.stdout.flush()
        self.tst_dataframe = data
        print()
        # print(self.tst_dataframe)
    
    
    def remove_X_pinmux_pads_in_dataframe(self):
        # get mux0 pads of pinmux in dataframe
        mux0_list_in_this_tst = []
        for col in self.tst_dataframe.columns:
            for pad in self.mux_group:
                if col == pad:
                    mux0_list_in_this_tst.append(col)
        
        # get all pads of pinmux in dataframe
        for mux0 in mux0_list_in_this_tst:
            if([x for x in self.mux_group[mux0] if x in self.tst_dataframe.columns]):
                self.all_mux_pads_in_this_tst.extend([x for x in self.mux_group[mux0] if x in self.tst_dataframe.columns])
        print(" "*SPACE_NUMBER + "All Mux Pad Number: ", len(self.all_mux_pads_in_this_tst))
        
        for pad, index in zip(self.all_mux_pads_in_this_tst, range(len(self.all_mux_pads_in_this_tst))):
            # print(pad, self.tst_dataframe[pad].unique())
            if(list(self.tst_dataframe[pad].unique()) == ['X']):
                # print("{:<30}".format(pad), self.tst_dataframe[pad].unique())
                self.tst_dataframe=self.tst_dataframe.drop(pad, axis=1)
                self.mux_X_pads_in_this_tst.append(pad)
            progress = index/(len(self.all_mux_pads_in_this_tst)-1)*100
            show_progress_percent(" "*SPACE_NUMBER + "Remove don't care mux pin in tst -> " , progress)
        print()
        print(" "*SPACE_NUMBER + "Don't Care Mux Pad Number:", len(self.mux_X_pads_in_this_tst))
        print(" "*SPACE_NUMBER + "Total ASSIGN Pads - Don't Care Mux Pad Number = ", len(self.tst_dataframe.columns))
        
        
        mux_pads_with_multi_alias = [mux_pad for mux_pad in self.all_mux_pads_in_this_tst if mux_pad not in self.mux_X_pads_in_this_tst]
        
        MUX_with_multi_alias_excel = create_excel('MUX_with_multi_alias', 'TST_Name', OUTPUT_FOLDER)
        MUX_with_multi_alias_excel_WB = openpyxl.load_workbook(MUX_with_multi_alias_excel)
        TST_Name_WS = MUX_with_multi_alias_excel_WB['TST_Name']
        
        multi_alias_flag = False
        intersection_list = {}
        mux_pads_with_multi_alias_summary = []
        for mux0_pad in self.mux_group:
            intersection_list[mux0_pad] = list(set(mux_pads_with_multi_alias) & set(self.mux_group[mux0_pad]))
            if (len(intersection_list[mux0_pad]) > 1):
                multi_alias_flag = True
        
        if(multi_alias_flag):
            global file_index
            file_index += 1
            MUX_with_multi_alias_excel_WS = MUX_with_multi_alias_excel_WB.create_sheet(str(file_index))       
            tst_name_font = Font(color='0000FF')
            TST_Name_WS.cell(file_index, 1).value = file_index
            TST_Name_WS.cell(file_index, 2).value = self.tst_name
            TST_Name_WS.cell(file_index, 2).hyperlink = f"#{MUX_with_multi_alias_excel_WS.title}!A1"
            TST_Name_WS.cell(file_index, 2).font = tst_name_font
            
            
            MUX_with_multi_alias_excel_WS.append([self.tst_name])
            MUX_with_multi_alias_excel_WS.append(['MUX0', 'All co-exist mux pads'])
            for mux0 in intersection_list:
                mux_pads_with_multi_alias_summary = []
                mux_pads_with_multi_alias_summary.append('Group ' + mux0)
                for mux_pad in mux_pads_with_multi_alias:
                    if mux_pad in intersection_list[mux0]:
                        mux_pads_with_multi_alias_summary.append(mux_pad)
                MUX_with_multi_alias_excel_WS.append(mux_pads_with_multi_alias_summary)
                        
            MUX_with_multi_alias_excel_WS.cell(1, 1).hyperlink = f"#{TST_Name_WS.title}!A{file_index}"
            MUX_with_multi_alias_excel_WS.cell(1, 1).font = tst_name_font
                
            MUX_with_multi_alias_excel_WB.save(MUX_with_multi_alias_excel)
        
        
        self.pads_not_exist_in_excel_but_in_tst = [x for x in self.tst_dataframe.columns if x not in self.all_digital_pads]
        print(" "*SPACE_NUMBER + "Pads not exist in excel but in tst ->", self.pads_not_exist_in_excel_but_in_tst)
        
        
        redundant_pads = [self.tstfile.split('/')[-1]]
        for col in self.tst_dataframe.columns:
            if col in self.pads_not_exist_in_excel_but_in_tst:
                redundant_pads.append(col)
                self.tst_dataframe = self.tst_dataframe.drop(col, axis=1)
        print(" "*SPACE_NUMBER + "Length of Alias each row after removing unused pads in tst now:", len(self.tst_dataframe.columns))
        
        redundant_pads_excel=create_excel('redundant_pads', 'redundant_pads', OUTPUT_FOLDER)
        redundant_wb = openpyxl.load_workbook(redundant_pads_excel)
        redundant_ws = redundant_wb['redundant_pads']
        redundant_ws.append(redundant_pads)
        redundant_wb.save(redundant_pads_excel)
        

    def write_pinmux_cleaned_tst(self):
        # self.tst_dataframe.to_excel(OUTPUT_FOLDER+"pinmux_cleaned.xlsx")
        output_tst = str(self.tstfile).replace(f"{INPUT_FOLDER}", OUTPUT_FOLDER)
        first_col = self.tst_dataframe.columns[0]
        for row in range(1, len(self.tst_dataframe.loc[:, first_col])+1):
            new_alias = ''
            for col in self.tst_dataframe.columns:
                new_alias += self.tst_dataframe.loc[row, col]
            self.new_alias[' /* '+ str(row) + ' */\n'] = new_alias
            
            progress = (row/(len(self.tst_dataframe.loc[:, first_col])))*100
            sys.stdout.write("\r" + " "*SPACE_NUMBER + "Build new tst alias content -> " + "{:.2f}".format(progress) +"%")
            sys.stdout.flush()
        print()
            
        # dump_dic(self.new_alias)
        all_need_remove_pads = []
        all_need_remove_pads.extend(self.mux_X_pads_in_this_tst)
        all_need_remove_pads.extend(self.pads_not_exist_in_excel_but_in_tst)
        
        assign_detect_flag = False
        for line, line_index in zip(self.tst_content, range(len(self.tst_content))):
            if(re.match(r"^ASSIGN.+,\n", line) and not assign_detect_flag):
                assign_detect_flag = True
                for mux_pad in all_need_remove_pads:
                    if(re.search(rf"{mux_pad}", line)):
                        line=line.replace(mux_pad+',', "")
                        line=line.replace(' '*16+'\n', "")
                self.tst_content[line_index] = line
            if(re.match(r"^\s+PAD_.+,\n", line) and assign_detect_flag):
                for mux_pad in all_need_remove_pads:
                    if(re.search(rf"{mux_pad}", line)):
                        line=line.replace(mux_pad+',', "")
                        line=line.replace(' '*16+'\n', "")
                self.tst_content[line_index] = line
            if(re.match(r"^\s+PAD_.+;\n", line) and assign_detect_flag):
                for mux_pad in all_need_remove_pads:
                    if(re.search(rf"{mux_pad}", line)):
                        line=line.replace(mux_pad+',', "")
                        line=line.replace(mux_pad+';', "")
                        line=line.replace(' '*16+'\n', "")
                        line=line.replace(',\n', ";\n")
                self.tst_content[line_index] = line
                # break
            if(re.match(r"\w+;\s\/\*\s\d+\s\*\/", line)):
                each_row_vector_num = line.split(';')[1]                
                self.tst_content[line_index] = self.new_alias[each_row_vector_num] + ';' + each_row_vector_num
                # break
            
            progress = (line_index/(len(self.tst_content)-1))*100
            sys.stdout.write("\r" + " "*SPACE_NUMBER + "Write new tst content -> " + "{:.2f}".format(progress) +"%")
            sys.stdout.flush()
        print()
        if self.extention == '.gz':
            self.tst_RW.write_gz_lines(output_tst, self.tst_content, 'wt')
        elif self.extention == '.tst':
            self.tst_RW.write_file_lines(output_tst, self.tst_content)


def main():
    print('='*50 + " Processing Start " + '='*50)
    start = time.time()
    
    CreateLoopOutputFolder(INPUT_FOLDER, OUTPUT_FOLDER)
    paths,             files             = Search_files(INPUT_FOLDER, ['.gz', '.tst'])
    channel_map_paths, channel_map_files = Search_files(INPUT_FOLDER, ['.xlsx'])
    
    if(len(channel_map_files) > 1):
        sys.exit("Error: There should be only 1 channel map excel file, please check!")
    
    global file_index
    file_index = 0
    for path, file in zip(paths, files):
        print(">>> " + file)
        CleanTstMuxPin(path + '/' + file, channel_map_paths[0] + '/' + channel_map_files[0], os.path.splitext(file)[1])
    
    # CleanTstMuxPin('./input/PA/ATPG_STUCK_PA_comp_capture_EDT_0P85V_0_20250831_v0.tst.gz', channel_map_paths[0] + '/' + channel_map_files[0])
    
    end = time.time()
    result=(divmod(end-start, 60))
    
    print(f"execuation time: {result[0]} m " + '{:.2f}'.format(result[1]) + "s")
    print('='*50 + " Processing Completed " + '='*50)

if __name__ == '__main__':
    main()