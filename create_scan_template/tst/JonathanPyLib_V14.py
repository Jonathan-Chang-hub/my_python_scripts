import os
import re
import openpyxl
import glob
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
import numpy as np
import gzip
import collections
import sys
import pandas as pd

INPUT_FOLDER:str = './input/'
OUTPUT_FOLDER:str = './output/'
SPACE_NUMBER:int = 4


def Search_files(folder_path, FileExt:list):
    file_path  = []
    source_files = []
    # 遍历指定文件夹及其子文件夹
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            # 检查文件扩展名
            for ext in FileExt:
                if file.endswith(ext):
                    # excel_files.append(os.path.join(root, file))
                    file_path.append(root)
                    source_files.append(file)
    return file_path, source_files


def GlobWantFileExt(FileExt) :
    Dmx_All_log_Array = []
    for All_log_file in glob.glob('*'+FileExt):
        print('The current folder file is = %s'%All_log_file)
        Dmx_All_log_Array.append(All_log_file)
    return(Dmx_All_log_Array)


def create_outputpath(output_folder, output_file_name, file_ext):
    currentDateAndTime = datetime.now()
    now = currentDateAndTime.strftime("%Y%m%d_%H%M%S")

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    OutputPath = output_folder+'/'+output_file_name+'_'+now+file_ext
    
    return OutputPath


def CreateLoopOutputFolder(Input_path:str, Output_path:str):
    
    if not os.path.exists(Output_path):
        os.mkdir(Output_path)
    
    temp_list = os.listdir(Input_path)
    
    for temp_list_each in temp_list:
        if os.path.isfile(Input_path + "/" + temp_list_each):
            continue
        else:
            if not os.path.exists(Output_path):
                os.mkdir(Output_path)
            CreateLoopOutputFolder(Input_path + "/" + temp_list_each, Output_path + "/" + temp_list_each)  


# //// read, write files ////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
class ReadWriteFiles:
    read_content = ""
    encode = 'utf-8'
    def read_file_lines(self, FileOpenPath):
        with open(FileOpenPath, 'r', encoding=self.encode) as fr:
            self.read_content = fr.readlines()
        fr.close()
    def write_file_lines(self, FileWritePath, WritedContentLine:list):
        with open(FileWritePath, 'w', encoding=self.encode) as fw:
            for k in range(len(WritedContentLine)):
                fw.write(WritedContentLine[k])
        fw.close()
    def read_file(self, FileOpenPath):
        with open(FileOpenPath, 'r', encoding=self.encode) as fr:
            self.read_content = fr.read()
        fr.close()
    def write_file(self, FileWritePath, WritedContent):
        with open(FileWritePath, 'w', encoding=self.encode) as fw:
            fw.write(WritedContent)
        fw.close()
    def append_file(self, FileWritePath, WritedContent):
        with open(FileWritePath, 'a', encoding=self.encode) as fa:
            fa.write(WritedContent)
        fa.close()
    def read_gz_lines(self, FileOpenPath, mode):
        with gzip.open(FileOpenPath, mode, encoding=self.encode) as f:
            self.read_content = f.readlines()
        f.close()
    def write_gz_lines(self, FileWritePath, WritedContent, mode = 'wt'):
        with gzip.open(FileWritePath, mode, encoding=self.encode) as f:
            for k in range(len(WritedContent)):
                f.write(WritedContent[k])
        f.close()
    def read_gz(self, FileOpenPath, mode):
        with gzip.open(FileOpenPath, mode, encoding=self.encode) as f:
            self.read_content = f.read()
        f.close()
            

def FindReMultiline(re_get_code_block, Content):
    match_obj = re.compile(re_get_code_block, re.DOTALL | re.MULTILINE)
    MatchedStrList = match_obj.findall(Content)
    return MatchedStrList


def rename_files(directory, oldname, newname):
   try:
       # 獲取目錄中的所有檔案名稱
       files = os.listdir(directory)
       # 遍歷所有檔案
       for file_name in files:
           # 檢查檔案名稱中是否包含 "oldname"
           if oldname in file_name:
               # 新檔案名稱
               new_name = file_name.replace(oldname, newname)
               # 完整路徑
               old_file_path = os.path.join(directory, file_name)
               new_file_path = os.path.join(directory, new_name)
               # 重命名檔案
               os.rename(old_file_path, new_file_path)
               print(f'Renamed: {file_name} -> {new_name}')
   except Exception as e:
       print(f"An error occurred: {e}")


def Transfer_StrLine_to_2DNumpy(Func_name_index, Func_end_index, initial_ptn1, FileContent):
    mos_lng_block = 0
    for i in range(len(Func_name_index)):
        if(Func_end_index[i]-Func_name_index[i]+1) > mos_lng_block:
            mos_lng_block = Func_end_index[i]-Func_name_index[i]+1

    FuncBlock=np.full((len(Func_name_index), mos_lng_block), initial_ptn1)

    for j in range(len(Func_name_index)):
        for k in range(Func_name_index[j]-1, Func_end_index[j]):
            FuncBlock[j][k+1-Func_name_index[j]] = FileContent[k]

    for j in range(len(Func_name_index)):
        for k in range(mos_lng_block):            
            if FuncBlock[j][k] == initial_ptn1:
                FuncBlock[j][k] = ''
    return FuncBlock, mos_lng_block


def find_specified_element_iloc_in_dataframe(dataframe, elements: list, dlog: bool):
    SymbolDic = {}
    for i in range(dataframe.shape[0]):
        for j in range(dataframe.shape[1]):
            if [x for x in elements if dataframe.iloc[i,j]==x]:
                SymbolDic[dataframe.iloc[i,j]] = [i,j]
                if dlog:
                    print(f"\"{dataframe.iloc[i,j]}\""+' (row,column): '+f"{i,j}")
    return SymbolDic


def IsFloatNum(str: str):
    s=str.split('.')
    if len(s)>2:
        return False
    else:
        for si in s:
            if not si.isdigit():
                return False
    return True


def txt_to_excel(InputFile, OutputFolder):
    txtname = str(InputFile).split('/')[-1]
    if os.path.splitext(txtname)[-1] == ".txt":
        #讀取 txt 檔案：防止讀取錯誤，讀取時需要指定編碼
        fopen = open(InputFile, 'r',encoding='utf-8')
        lines = fopen.readlines()
        #寫入 excel表
        file = openpyxl.Workbook()
        sheet = file.active
        # 新建一個sheet
        sheet.title = txtname.split('/')[-1].strip(".txt")
        # sheet.title = txtname.strip(fileExt)
        
        i = 0
        for line in lines:
            # strip 移出字串頭尾的換行
            # line = line.strip('\n')
            
            # 用','替換掉'\t',很多行都有這個問題，導致不能正確把各個特徵值分開
            line = line.replace("\t", ",")
            line = line.replace("=", "\'=")
            line = line.replace('	', ',')
            # print(line)
            line = line.split(',')
            # 一共7個欄位
            for index in range(len(line)):
                sheet.cell(i+1, index+1, line[index])
            # 行數遞增
            i = i + 1
        if not os.path.exists(OutputFolder):
            os.mkdir(OutputFolder)
        if(InputFile==OutputFolder):
            file.save(f'{InputFile}'.replace('txt', 'xlsx'))
        else:
            file.save(f'{OutputFolder}'+'/'+txtname.split('/')[-1].replace('txt', 'xlsx'))


def unzip_gz(file_name, outputfolder, UnzipAndWrite:bool=False ,otheroutput:bool=False):
    f_name = str(file_name).replace(".gz", "")
    g_file = gzip.GzipFile(file_name)
    g_file_content=g_file.read()
    if UnzipAndWrite:
        file=f_name.split("/")[-1]
        print("Unzip-->"+str(file))
        if otheroutput:
            if not os.path.exists(outputfolder):
                os.makedirs(outputfolder)
            open(outputfolder+'/'+file, "wb+").write(g_file_content)
        else:
            open(f_name, "wb+").write(g_file_content)
    g_file.close()
    return g_file_content


def unique_list(List: list):
    folderListSet=set(List)
    # print(folderListSet)
    uniqueList = list(folderListSet)
    # print(uniquefolderlist)
    return uniqueList


def makeNestedStruct(n, type):
    if n == 1: return collections.defaultdict(type)
    else:      return collections.defaultdict(lambda: makeNestedStruct(n-1, type))


def create_excel(ExcelName:str, ExcelSheet:str, folder:str):
    NewExcel = folder+'/'+ExcelName+'.xlsx'
    if(not os.path.exists(NewExcel)):
        SummaryWb = openpyxl.Workbook(NewExcel)
        SummaryWb.create_sheet(ExcelSheet)
        SummaryWb.save(NewExcel)
    return NewExcel


def dump_dic(dictionary:dict):
    print('='*80+' Dump Dictionary '+'='*80)
    for key, val in dictionary.items():
        print(key, "->", val)
    print('='*80+' Dump Dictionary End '+'='*80)
    

def clear_screen():
    if os.name == 'nt':
        os.system('cls')
    else:
        os.system('clear')
    

def show_progress_percent(messages:str , progress:float):
    sys.stdout.write(f"\r{messages}{progress :>.2%}")
    sys.stdout.flush()


def write_dataframe_cols_into_rows(df:pd.DataFrame):
    with open('columns_values_only.txt', 'w') as f:
        for column in df.columns:
            values = df[column].astype(str).tolist()
            f.write(' '.join(values) + '\n')
    f.close()