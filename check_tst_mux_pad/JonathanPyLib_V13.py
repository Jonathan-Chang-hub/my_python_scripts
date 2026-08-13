import os
import re
import openpyxl
import glob
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
import numpy as np
import gzip
import collections
import sys



def Search_files(folder_path, FileExt:list):
    file_path  = []
    source_files = []
    # 遍历指定文件夹及其子文件夹
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            # 检查文件扩展名
            for ext in FileExt:
                if file.endswith(ext):
                    # excel_files.append(os.path.join(root, file))
                    file_path.append(root)
                    source_files.append(file)
    return file_path, source_files


def GlobWantFileExt(FileExt) :
    Dmx_All_log_Array = []
    for All_log_file in glob.glob('*'+FileExt):
        print('The current folder file is = %s'%All_log_file)
        Dmx_All_log_Array.append(All_log_file)
    return(Dmx_All_log_Array)


def read_xlsm_data(file_path, title_row):
    # 打开 Excel 文件
    wb = openpyxl.load_workbook(file_path, data_only = 'True')

    # 用于存储所有工作表中的数据
    Test_Name = []
    Name = []

    # 遍历所有工作表
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # 查找标题行
        for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value == 'Test Instances':  # 替换为你要查找的标题
                    for row2 in sheet.iter_rows(min_row=title_row, max_row=title_row):
                        for cell2 in row2:
                            if cell2.value == 'Test Name':  # 替换为你要查找的标题
                                # 获取标题所在的列号
                                title_column_number = cell2.column_letter
                                title_column = column_index_from_string(title_column_number)
                                title_row_data = cell2.row

                                # 读取标题列下的数据
                                for row in sheet.iter_rows(min_row=title_row_data + 1, min_col=title_column, max_col=title_column):
                                    for cell in row:
                                        if cell.value != None :
                                            #寫入陣列
                                            Test_Name.append(cell.value)

                                # 存储数据到字典中
                                # Test_Name.append(data)


                            if cell2.value == 'Name':  # 替换为你要查找的标题
                                # 获取标题所在的列号
                                title_column_number = cell2.column_letter
                                title_column = column_index_from_string(title_column_number)
                                title_row_data = cell2.row

                                # 读取标题列下的数据
                                for row in sheet.iter_rows(min_row=title_row_data + 1, min_col=title_column, max_col=title_column):
                                    for cell in row:
                                        if cell.value != None :
                                            #寫入陣列
                                            Name.append(cell.value)

                                # 存储数据到字典中
                                # Name.append(data)
    Instances_VBA = {}
    for i in range(len(Name)):
        Instances_VBA[Test_Name[i]] = Name[i]

    return Instances_VBA


def create_outputpath(output_folder, output_file_name, file_ext):
    currentDateAndTime = datetime.now()
    now = currentDateAndTime.strftime("%Y%m%d_%H%M%S")

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    OutputPath = output_folder+'/'+output_file_name+'_'+now+file_ext
    
    return OutputPath


def CreateLoopOutputFolder(Input_path, Output_path):
    
    if not os.path.exists(Output_path):
        os.mkdir(Output_path)
    
    temp_list = os.listdir(Input_path)
    
    for temp_list_each in temp_list:
        if os.path.isfile(Input_path + "/" + temp_list_each):
            continue
        else:
            if not os.path.exists(Output_path):
                os.mkdir(Output_path)
            CreateLoopOutputFolder(Input_path + "/" + temp_list_each, Output_path + "/" + temp_list_each)  


#抓取txt Instances
def grab_txt_instances():
    Instance_txt = GlobWantFileExt('.txt')
    Instance_list =[]
    with open('./'+str(Instance_txt[0])) as f :
        lines = f.readlines()
        for line in lines :
            a = line.split('	')
            # a = line.split(' ')
            # if a[1] =
            if len(a) > 9 :
                if a[3] !='' :
                    Instance_list.append(a[3])
    return Instance_list


def export_instances_output(outputPMICTemp, Instance_list, Instance_TestName_VBA_Database):
    wea = openpyxl.Workbook(outputPMICTemp)
    s1 = wea.create_sheet('Home')
    wea.save(outputPMICTemp)

    Type_name3 = ['Flow Step', 'VBA_Name', 'sheet_name','Owner', 'Total Test time(ms)', 'Total Wait Time(ms)']
    ws = openpyxl.load_workbook(outputPMICTemp, data_only = 'True')
    s1 =ws['Home']
    s1.append(Type_name3)

    for i in Instance_list :
        if(i in Instance_TestName_VBA_Database) :
            a = [i,Instance_TestName_VBA_Database[i]]
            s1.append(a)
        else :
            continue
            # a = [i,'NA']
            # s1.append(a)
        ws.save(outputPMICTemp)

# search VBA code title name inside folder path and print them into excel
def export_VBcode_to_sheet(outputPMICTemp, Instance_list, VBT_path, VBT_files, VBA_Database, checkfilenum, UsedBASfile):
    
    global FileName_split_to_sheet
    FileName_split_to_sheet = ""
    
    wea = openpyxl.load_workbook(outputPMICTemp, data_only = 'True')
    s1 = wea.create_sheet('bas_function_name')
    wea.save(outputPMICTemp)
    
    Type_name = ['Code', 'Test time(ms)', 'Wait Time(ms)', 'Total Test time(ms)', 'Total Wait Time(ms)']
    Type_name2 = ['Path','file_name','VBA_Name']
    
    ws = openpyxl.load_workbook(outputPMICTemp, data_only = 'True')
    s1 =ws['bas_function_name']
    s1.append(Type_name2)
    WsHome =ws['Home']
    
    # Instance_list.remove("PreJob Reset")
    for i in Instance_list:
        if(i in VBA_Database):
            for j in range(len(VBT_files)) : 
                file_path = VBT_path[j] + '/' + VBT_files[j]
                file_name_split = VBT_files[j].split('.')
                if(VBA_Database[i] == file_name_split[0]):
                    UsedBASfile += 1
                    with open(file_path, 'r', encoding="windows-1252") as f :
                        wordf = []
                        wordf2 = []
                        lines = f.readlines()
                        
                        # avoid the sheet name characters more than 32
                        if len(file_name_split[0])>31 :
                            FileName_split_to_sheet = str(file_name_split[0])[0:20]+'_'+str(j)
                        else:
                            FileName_split_to_sheet = str(file_name_split[0])

                        # sumarize the source file path and list the corresponding VBT code 
                        wordf = [str(VBT_path[j]), str(FileName_split_to_sheet), file_name_split[0]]
                        s1.append(wordf)
                        WsHome.cell(UsedBASfile,3).value = str(FileName_split_to_sheet)
                        # find the corresponding VBT code in assigned path and put them into excel sheet
                        checkfilenum += 1
                        print(str(checkfilenum)+" Sheet:"+str(FileName_split_to_sheet)+", VBT:"+str(file_name_split[0]))
                        s2 = ws.create_sheet(FileName_split_to_sheet)
                        s2.append(['Home'])
                        s2.append([i])
                        s2.append(Type_name)
                        for line in lines : 
                            wordf2 = [line]
                            s2.append(wordf2)
                        ws.save(outputPMICTemp)
                    f.close()                    
    return checkfilenum


def import_link_to_excel(file_path, title_row, checkfilenum):
    # 打开 Excel 文件
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # 用于存储所有工作表中的数据
    Name = []

    # 遍历所有工作表
    sheet = wb['Home']
    for row in sheet.iter_rows(min_row=title_row, max_row=title_row):
        for cell in row:
            if cell.value == 'sheet_name':  # 替换为你要查找的标题
                # 获取标题所在的列号
                title_column_number = cell.column_letter
                title_column = column_index_from_string(title_column_number)
                
                # 读取标题列下的数据
                data = []
                for row in sheet.iter_rows(min_row=title_row + 1, min_col=title_column, max_col=title_column):
                    for cell in row:
                        if len(cell.value) <= 32 :
                            w2 = wb[str(cell.value)]
                            # sheet.cell(row=cell.row, column=cell.column).hyperlink = f"#{w2.title}!A1"
                            sheet.cell(row=cell.row, column=cell.column-2).hyperlink = f"#{w2.title}!A1"
                            w2.cell(row=1, column=1).hyperlink = f"#{sheet.title}!"+get_column_letter(cell.column-2)+str(cell.row)
                        else :
                            Name.append(cell)
    wb.save(file_path)
    
    wb = openpyxl.load_workbook(file_path)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if sheet_name != 'Home' and sheet_name != 'bas_function_name':
        # shoot = wb['Home']
        # 查找标题行
            for row in sheet.iter_rows(min_row=2, max_row=2):
                for cell in row:
                    sheet.cell(row=4, column=4, value="=IF(COUNTIF(B4:B2000,\"NA\")>0,\"NA\",SUM(B4:B2000))")
                    sheet.cell(row=4, column=5, value="=SUM($C$4:$C$2000)")
        elif sheet_name == 'Home':
            row_index = 1
            for row2 in sheet.iter_rows(min_row=2, max_row=checkfilenum+1, min_col=3, max_col=3):
                for cell in row2:
                    row_index += 1
                    sheet.cell(row=row_index, column=5, value="="+str(cell.value)+"!D4")
                    sheet.cell(row=row_index, column=6, value="="+str(cell.value)+"!E4")
    wb.save(file_path)


def import_time_to_excel(file_path, title_row, re_list):
    # 打开 Excel 文件
    wb = openpyxl.load_workbook(file_path)
    
    # 用于存储所有工作表中的数据
    Attribute = []

    
    # 遍历所有工作表
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if sheet_name != 'Home' and sheet_name != 'bas_function_name':
            print(sheet.title)
            # 查找标题行
            for row in sheet.iter_rows(min_row=title_row, max_row=title_row):
                if row != 'Home' :
                    for cell in row:
                        if cell.value == 'Code':  # 替换为你要查找的标题
                            # 获取标题所在的列号
                            title_column_number = cell.column_letter
                            title_column = column_index_from_string(title_column_number)
                            
                            # 读取标题列下的数据
                            data = []
                            for row in sheet.iter_rows(min_row=title_row + 1, min_col=title_column, max_col=title_column):
                                for cell in row:
                                    for key, value in re_list.items() :
                                        if cell.value is not None:
                                            result = re.match(fr'{key}', cell.value)
                                            result1 = re.findall(r'[^\']+[ ]*thehdw.Wait\b\s(\d+).*', cell.value)
                                            result3 = re.findall(r'[^\']+[ ]*thehdw.Wait\b\s(\d.\d+)', cell.value)
                                            if result :
                                                sheet.cell(row=cell.row, column=cell.column+1, value=value[0])
                                                sheet.cell(row=cell.row, column=cell.column+2, value=value[1])
                                            elif result3 :
                                                sheet.cell(row=cell.row, column=cell.column+1, value=float(result3[0]))
                                                sheet.cell(row=cell.row, column=cell.column+2, value=float(result3[0]))
                                            elif result1 :
                                                
                                                sheet.cell(row=cell.row, column=cell.column+1, value=int(result1[0]))
                                                sheet.cell(row=cell.row, column=cell.column+2, value=int(result1[0]))
                                            else :
                                                result2 = re.match(r'[^\']+[ ]*thehdw.Wait\b.*', cell.value)
                                                if result2 :
                                                    sheet.cell(row=cell.row, column=cell.column+1, value=1)
                                                    sheet.cell(row=cell.row, column=cell.column+2, value=1)
                wb.save(file_path)


                            # 存储数据到字典中
                            # Attribute[sheet_name] = data
                            # print('Attribute')
                                
    return Attribute


def read_regular_xlsx(file_path):
    # 打开 Excel 文件
    wb = openpyxl.load_workbook(file_path)
   
    # 用于存储所有工作表中的数据
    Attribute = {}
   
    # 遍历所有工作表
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
       
        # 查找标题行
        for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value == 'python regular':  # 替换为你要查找的标题
                    # 获取标题所在的列号
                    title_column_number = cell.column_letter
                    title_column = column_index_from_string(title_column_number)
                   
                    # 读取标题列下的数据
                    data = []
                    for row in sheet.iter_rows(min_row=2, min_col=title_column, max_col=title_column):
                        for cell in row:
                            data.append(cell.value)
                            time = sheet.cell(row = cell.row, column = cell.column+1)
                            Wait_time = sheet.cell(row = cell.row, column = cell.column+2)
                            Attribute[cell.value] = [time.value, Wait_time.value]
                            # print(Attribute[cell.value])
                        
        # print(len(data))
    return Attribute


# //// read, write files ////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////
class ReadWriteFiles:
    read_content = ""
    encode = 'utf-8'
    def read_file_lines(self, FileOpenPath):
        with open(FileOpenPath, 'r', encoding=self.encode) as fr:
            self.read_content = fr.readlines()
        fr.close()
    def write_file_lines(self, FileWritePath, WritedContentLine:list):
        with open(FileWritePath, 'w', encoding=self.encode) as fw:
            for k in range(len(WritedContentLine)):
                fw.write(WritedContentLine[k])
        fw.close()
    def read_file(self, FileOpenPath):
        with open(FileOpenPath, 'r', encoding=self.encode) as fr:
            self.read_content = fr.read()
        fr.close()
    def write_file(self, FileWritePath, WritedContent):
        with open(FileWritePath, 'w', encoding=self.encode) as fw:
            fw.write(WritedContent)
        fw.close()
    def append_file(self, FileWritePath, WritedContent):
        with open(FileWritePath, 'a', encoding=self.encode) as fa:
            fa.write(WritedContent)
        fa.close()
    def read_gz_lines(self, FileOpenPath, mode):
        with gzip.open(FileOpenPath, mode, encoding=self.encode) as f:
            self.read_content = f.readlines()
        f.close()
    def write_gz_lines(self, FileWritePath, WritedContent, mode = 'wt'):
        with gzip.open(FileWritePath, mode, encoding=self.encode) as f:
            for k in range(len(WritedContent)):
                f.write(WritedContent[k])
        f.close()
    def read_gz(self, FileOpenPath, mode):
        with gzip.open(FileOpenPath, mode, encoding=self.encode) as f:
            self.read_content = f.read()
        f.close()
            

def FindReMultiline(re_get_code_block, Content):
    match_obj = re.compile(re_get_code_block, re.DOTALL | re.MULTILINE)
    MatchedStrList = match_obj.findall(Content)
    return MatchedStrList


def rename_files(directory, oldname, newname):
   try:
       # 獲取目錄中的所有檔案名稱
       files = os.listdir(directory)
       # 遍歷所有檔案
       for file_name in files:
           # 檢查檔案名稱中是否包含 "oldname"
           if oldname in file_name:
               # 新檔案名稱
               new_name = file_name.replace(oldname, newname)
               # 完整路徑
               old_file_path = os.path.join(directory, file_name)
               new_file_path = os.path.join(directory, new_name)
               # 重命名檔案
               os.rename(old_file_path, new_file_path)
               print(f'Renamed: {file_name} -> {new_name}')
   except Exception as e:
       print(f"An error occurred: {e}")


def Transfer_StrLine_to_2DNumpy(Func_name_index, Func_end_index, initial_ptn1, FileContent):
    mos_lng_block = 0
    for i in range(len(Func_name_index)):
        if(Func_end_index[i]-Func_name_index[i]+1) > mos_lng_block:
            mos_lng_block = Func_end_index[i]-Func_name_index[i]+1

    FuncBlock=np.full((len(Func_name_index), mos_lng_block), initial_ptn1)

    for j in range(len(Func_name_index)):
        for k in range(Func_name_index[j]-1, Func_end_index[j]):
            FuncBlock[j][k+1-Func_name_index[j]] = FileContent[k]

    for j in range(len(Func_name_index)):
        for k in range(mos_lng_block):            
            if FuncBlock[j][k] == initial_ptn1:
                FuncBlock[j][k] = ''
    return FuncBlock, mos_lng_block


def find_specified_element_iloc_in_dataframe(dataframe, elements: list, dlog: bool):
    SymbolDic = {}
    for i in range(dataframe.shape[0]):
        for j in range(dataframe.shape[1]):
            if [x for x in elements if dataframe.iloc[i,j]==x]:
                SymbolDic[dataframe.iloc[i,j]] = [i,j]
                if dlog:
                    print(f"\"{dataframe.iloc[i,j]}\""+' (row,column): '+f"{i,j}")
    return SymbolDic


def IsFloatNum(str: str):
    s=str.split('.')
    if len(s)>2:
        return False
    else:
        for si in s:
            if not si.isdigit():
                return False
    return True


def txt_to_excel(InputFile, OutputFolder):
    txtname = str(InputFile).split('/')[-1]
    if os.path.splitext(txtname)[-1] == ".txt":
        #讀取 txt 檔案：防止讀取錯誤，讀取時需要指定編碼
        fopen = open(InputFile, 'r',encoding='utf-8')
        lines = fopen.readlines()
        #寫入 excel表
        file = openpyxl.Workbook()
        sheet = file.active
        # 新建一個sheet
        sheet.title = txtname.split('/')[-1].strip(".txt")
        # sheet.title = txtname.strip(fileExt)
        
        i = 0
        for line in lines:
            # strip 移出字串頭尾的換行
            # line = line.strip('\n')
            
            # 用','替換掉'\t',很多行都有這個問題，導致不能正確把各個特徵值分開
            line = line.replace("\t", ",")
            line = line.replace("=", "\'=")
            line = line.replace('	', ',')
            # print(line)
            line = line.split(',')
            # 一共7個欄位
            for index in range(len(line)):
                sheet.cell(i+1, index+1, line[index])
            # 行數遞增
            i = i + 1
        if not os.path.exists(OutputFolder):
            os.mkdir(OutputFolder)
        if(InputFile==OutputFolder):
            file.save(f'{InputFile}'.replace('txt', 'xlsx'))
        else:
            file.save(f'{OutputFolder}'+'/'+txtname.split('/')[-1].replace('txt', 'xlsx'))


def unzip_gz(file_name, outputfolder, UnzipAndWrite:bool=False ,otheroutput:bool=False):
    f_name = str(file_name).replace(".gz", "")
    g_file = gzip.GzipFile(file_name)
    g_file_content=g_file.read()
    if UnzipAndWrite:
        file=f_name.split("/")[-1]
        print("Unzip-->"+str(file))
        if otheroutput:
            if not os.path.exists(outputfolder):
                os.makedirs(outputfolder)
            open(outputfolder+'/'+file, "wb+").write(g_file_content)
        else:
            open(f_name, "wb+").write(g_file_content)
    g_file.close()
    return g_file_content


def unique_list(List: list):
    folderListSet=set(List)
    # print(folderListSet)
    uniqueList = list(folderListSet)
    # print(uniquefolderlist)
    return uniqueList


def makeNestedStruct(n, type):
    if n == 1: return collections.defaultdict(type)
    else:      return collections.defaultdict(lambda: makeNestedStruct(n-1, type))


def create_excel(ExcelName:str, ExcelSheet:str, folder:str):
    NewExcel = folder+'/'+ExcelName+'.xlsx'
    if(not os.path.exists(NewExcel)):
        SummaryWb = openpyxl.Workbook(NewExcel)
        SummaryWb.create_sheet(ExcelSheet)
        SummaryWb.save(NewExcel)
    return NewExcel


def dump_dic(dictionary:dict):
    print('='*80+' Dump Dictionary '+'='*80)
    for key, val in dictionary.items():
        print(key, "->", val)
    print('='*80+' Dump Dictionary End '+'='*80)
    

def clear_screen():
    if os.name == 'nt':
        os.system('cls')
    else:
        os.system('clear')
        

def show_progress_percent(messages:str , progress:float):
    sys.stdout.write("\r"+ messages + "{:.2f}".format(progress) +"%")
    sys.stdout.flush()
