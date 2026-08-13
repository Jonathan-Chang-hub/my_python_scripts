import re
import time
import openpyxl
from openpyxl.styles import Font
import pandas as pd
from JonathanPyLib_V13 import *


INPUT_FOLDER = './input/'
OUTPUT_FOLDER = './output/'
SPACE_NUMBER:int = 4


class CheckTstMuxPad():
    
    def __init__(self, tstfile, pin_mux_alias_excel, extention):
        self.tstfile = tstfile
        self.tst_name = str(tstfile.split('/')[-1]).strip('.tst.gz')
        self.pin_mux_alias_excel =  pin_mux_alias_excel
        self.tst_RW = ReadWriteFiles()
        if extention == '.gz':
            self.tst_RW.read_gz_lines(self.tstfile, 'rt')       # read tst.gz or .tst file content
        elif extention == '.tst':
            self.tst_RW.read_file_lines(self.tstfile)
        self.tst_content = self.tst_RW.read_content
        
        self.mux_group = {}                                 # mux_group['PAD_BRCS0'] = ['PAD_BRCS0', 'PAD_MOD_TX_N00'...] from excel
        self.establish_pin_mux_group()                      # establish pin mux group dictionary from channel map
        
        self.all_digital_pads = []                          # all used digital pads in alias excel
        self.get_all_digital_pads()
        
        self.tst_assign_pads = []                           # ASSIGN pads in tst
        self.alias = {}                                     # self.alias['/* 1 */'] = ['X', 'X', '0', 'H', ...]
        self.extract_assign_pads_and_alias()
        
        self.check_redudant_and_pin_mux()
        
        
    def extract_assign_pads_and_alias(self):
        assign_detect_flag = False
        for line in self.tst_content:
            if(re.match(r"^ASSIGN.+,\n", line) and not assign_detect_flag):
                assign_detect_flag = True
                line=line.strip('ASSIGN').strip().strip(',')
                self.tst_assign_pads = line.split(',')
            if(re.match(r"^\s+PAD_.+,\n", line) and assign_detect_flag):
                line=line.strip().strip(',')
                self.tst_assign_pads.extend(line.split(','))
            if(re.match(r"^\s+PAD_.+;\n", line) and assign_detect_flag):
                line=line.strip().strip(';')
                self.tst_assign_pads.extend(line.split(','))
                break
        print(" "*SPACE_NUMBER + "Total ASSIGN pads unmber in tst: ", len(self.tst_assign_pads))

        
    def establish_pin_mux_group(self):
        pin_mux_df = pd.read_excel(self.pin_mux_alias_excel, sheet_name='ALIAS')
        for row in pin_mux_df.loc[:, 'MUX0']:
            if (str(row) != 'nan'):
                self.mux_group[row] = [str(row)]
        for i in range(len(pin_mux_df.loc[:, 'MUX0'])):
            for mux in pin_mux_df.columns:
                if(mux != 'MUX0' and re.match(r"MUX\d+", mux, re.IGNORECASE) and str(pin_mux_df.loc[i, mux]) != 'nan'):
                    self.mux_group[pin_mux_df.loc[i, 'MUX0']].append(str(pin_mux_df.loc[i, mux]))
        # dump_dic(self.mux_group)
    
    
    def get_all_digital_pads(self):
        all_digital_pads_df = pd.read_excel(self.pin_mux_alias_excel, sheet_name='ALIAS')
        for row in range(len(all_digital_pads_df.iloc[:, 0])):
            for col in range(len(all_digital_pads_df.columns)):
                if (str(all_digital_pads_df.iloc[row, col]) != 'nan'):
                    self.all_digital_pads.append(str(all_digital_pads_df.iloc[row, col]))
        print(" "*SPACE_NUMBER + "All assigned digital & mux pads number in \"" + str(self.pin_mux_alias_excel).split('/')[-1] +"\":", len(self.all_digital_pads))
        
    
    def check_redudant_and_pin_mux(self):
        co_exist_mux_excel = create_excel('co-exist_mux', 'TST_NAME', OUTPUT_FOLDER)
        co_exist_mux_excel_wb = openpyxl.load_workbook(co_exist_mux_excel)
        TST_Name_ws = co_exist_mux_excel_wb['TST_NAME']
        
        multi_alias_flag = False
        co_exist_mux_in_this_mux0 = {}
        for mux0 in self.mux_group:
            co_exist_mux_in_this_mux0[mux0] = []
            co_exist_mux_in_this_mux0[mux0] =  list(set(self.mux_group[mux0]) & set(self.tst_assign_pads))
            if len(co_exist_mux_in_this_mux0[mux0]) > 1:
                multi_alias_flag = True
                co_exist_mux_in_this_mux0[mux0].insert(0, 'Group ' + mux0)
        
        if multi_alias_flag:
            global file_index
            file_index += 1
            co_exist_mux_excel_wb.create_sheet(str(file_index))
            co_exist_mux_excel_ws = co_exist_mux_excel_wb[str(file_index)]
            tst_name_font = Font(color='0000FF')
            
            TST_Name_ws.append([str(file_index), self.tst_name])
            TST_Name_ws.cell(file_index, 1).value = file_index
            TST_Name_ws.cell(file_index, 2).value = self.tst_name
            TST_Name_ws.cell(file_index, 2).hyperlink = f"#{co_exist_mux_excel_ws.title}!A1"
            TST_Name_ws.cell(file_index, 2).font = tst_name_font
            
            co_exist_mux_excel_ws.append([self.tst_name])
            co_exist_mux_excel_ws.append(['MUX0', 'All co-exist mux pads'])
            co_exist_mux_excel_ws.cell(1, 1).hyperlink = f"#{TST_Name_ws.title}!A{file_index}"
            co_exist_mux_excel_ws.cell(1, 1).font = tst_name_font
            
            for mux0 in self.mux_group:
                co_exist_mux_excel_ws.append(co_exist_mux_in_this_mux0[mux0])
        
        co_exist_mux_excel_wb.save(co_exist_mux_excel)
        
        self.pads_not_exist_in_excel_but_in_tst = [x for x in self.tst_assign_pads if x not in self.all_digital_pads]
        print(" "*SPACE_NUMBER + "Pads not exist in excel but in tst ->", self.pads_not_exist_in_excel_but_in_tst)
        
        redundant_pads = [self.tstfile.split('/')[-1]]
        for pad in self.pads_not_exist_in_excel_but_in_tst:
            redundant_pads.append(pad)
        
        redundant_pads_excel=create_excel('pads_not_exist_in_excel_but_in_tst', 'redundant_pads', OUTPUT_FOLDER)
        redundant_wb = openpyxl.load_workbook(redundant_pads_excel)
        redundant_ws = redundant_wb['redundant_pads']
        redundant_ws.append(redundant_pads)
        redundant_wb.save(redundant_pads_excel)


def main():
    print('='*50 + " Processing Start " + '='*50)
    start = time.time()
    
    # CreateLoopOutputFolder(INPUT_FOLDER, OUTPUT_FOLDER)
    if not os.path.exists(OUTPUT_FOLDER):
        os.mkdir(OUTPUT_FOLDER)
    
    paths,             files             = Search_files(INPUT_FOLDER, ['.gz', '.tst'])
    channel_map_paths, channel_map_files = Search_files(INPUT_FOLDER, ['.xlsx'])
    
    if(len(channel_map_files) > 1):
        sys.exit("Error: There should be only 1 channel map excel file, please check!")
    
    global file_index
    file_index = 0
    for path, file in zip(paths, files):
        print(">>> " + file)
        CheckTstMuxPad(path + '/' + file, channel_map_paths[0] + '/' + channel_map_files[0], os.path.splitext(file)[1])
    
    end = time.time()
    result=(divmod(end-start, 60))
    
    print(f"execuation time: {result[0]} m " + '{:.2f}'.format(result[1]) + "s")
    print('='*50 + " Processing Completed " + '='*50)

if __name__ == '__main__':
    main()